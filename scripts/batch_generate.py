#!/usr/bin/env python3
"""
Batch Problem Generator for CodeCombat 2026

Reads questions from Deloitte_100_Coding_Questions.xlsx,
generates problems + 5 harnesses via NVIDIA NIM API (Nemotron),
and stores directly in PostgreSQL.

Usage:
  python3 batch_generate.py

Prerequisites:
  pip install openpyxl requests psycopg2
  Export: NVIDIA_API_KEY=...
"""

import openpyxl, requests, json, time, re, sys, os, subprocess, tempfile

# ─── Config ─────────────────────────────────────────────────────────────────
SHEET_PATH = "../Sheets/Deloitte_100_Coding_Questions.xlsx"
API_KEY = os.environ.get("NVIDIA_API_KEY", "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")
API_URL = "https://integrate.api.nvidia.com/v1/chat/completions"
MODEL   = "nvidia/nemotron-3-ultra-550b-a55b"  # or: deepseek-ai/deepseek-v4-flash

DB_HOST = os.environ.get("DB_HOST", "localhost")
DB_PORT = os.environ.get("DB_PORT", "5432")
DB_NAME = os.environ.get("DB_NAME", "codecombat")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASS = os.environ.get("DB_PASSWORD", "postgres")

# ─── Prompts ────────────────────────────────────────────────────────────────

PASS1_SYSTEM = r"""You are a problem designer for CodeCombat 2026, a competitive-programming judge.
The user gives EITHER a LeetCode problem (name or number) OR a custom problem
described in their own words (possibly a story). Produce ONE raw JSON object —
the problem spec — and NOTHING else.

REFRAMING (critical — read first):
The judge runs exactly ONE method that the candidate implements — exactly like a
LeetCode `Solution` class with a single public method. If the requested problem is
interactive or design/API-based you MUST REFRAME it into an equivalent single method:
pass the hidden state as explicit inputs and return a concrete value. `signature.name`
is that method's name and MUST be a clean camelCase verb phrase (e.g. twoSum,
reverseList, maxDepth).

DATA STRUCTURES (LeetCode-style):
  ListNode — serialized as [1,2,3]; empty = []
  TreeNode — level-order BFS with null, e.g. [1,null,2,3]; empty = []
Choose ListNode/TreeNode ONLY for problems genuinely about those structures.

Keep "description" concise (60-120 words).

TOPICS: Assign comma-separated from: Array, String, Two Pointers, Sliding Window,
Binary Search, Hash Table, Linked List, Stack, Queue, Tree, Binary Tree, BST,
Heap, Graph, Dynamic Programming, Greedy, Sorting, Bit Manipulation, Math,
Recursion, Backtracking, DFS, BFS, Union Find, Trie, Divide and Conquer, Simulation

Shape:
{
  "problem": {
    "title": "short title",
    "description": "concise 60-120 word statement",
    "inputFormat": "how input is described to solver",
    "outputFormat": "what to return/print",
    "constraints": "newline-separated",
    "timeLimit": 5,
    "memoryLimit": 256,
    "level": "EASY|MEDIUM|HARD",
    "topics": "Array, Hash Table",
    "example1": "Input: ...\nOutput: ...\nExplanation: ...",
    "example2": "...",
    "example3": "..."
  },
  "signature": {
    "name": "camelCaseName",
    "returnType": "int",
    "params": [{"name": "nums", "type": "int[]"}]
  },
  "referenceSolutionPython": "def funcName(params):\n    ...\n    return answer",
  "tests": [
    {"args":{"param":<val>},"expected":<val>,"hidden":false},
    ...EXACTLY 6 entries (first 3 hidden=false, last 3 hidden=true)
  ]
}

Allowed types: int, long, double, boolean, String, char,
int[], long[], double[], boolean[], String[], int[][], ListNode, TreeNode, ListNode[].

referenceSolutionPython must be CORRECT, executable Python. imports at top allowed.
For ListNode/TreeNode problems: receive serialized list, return serialized list.
6 test inputs with meaningful edge cases.

OUTPUT: only the JSON object. First char {, last char }. No markdown."""

HARNESS_SYSTEM = r"""You generate five CodeCombat 2026 judge harnesses (JAVA, CPP, PYTHON, JAVASCRIPT, C)
for a given problem spec. INPUT: a JSON spec with 'signature' and 'tests'.
OUTPUT: raw source code only.

Emit harnesses in this order, each preceded by EXACTLY:
===HARNESS:JAVA===
===HARNESS:CPP===
===HARNESS:PYTHON===
===HARNESS:JAVASCRIPT===
===HARNESS:C===
No markdown fences, no prose — just header lines and code.

Harness contract:
• One self-contained, runnable file. NO stdin.
• Markers: // USER_CODE_START / // USER_CODE_END (Python: # USER_CODE_START / # USER_CODE_END)
• LeetCode-style: JAVA/CPP/PYTHON → class Solution with ONE public method
  JAVASCRIPT/C → single top-level function
• CRITICAL: Every test() call in main() MUST be wrapped in try-catch so a crash
  in one test case prints FAIL and continues. Never let an unhandled exception
  abort the whole harness.
• test() prints EXACTLY one line per test case:
    visible pass:  TC:<n>:PASS
    hidden  pass:  TC:<n>:PASS:hidden
    visible fail:  TC:<n>:FAIL:input=<repr>:expected=<exp>:got=<got>
    hidden  fail:  TC:<n>:FAIL:hidden
• Hidden tests never print input/expected/got.
• Keep printed input/expected/got free of ':' characters.

JAVA: public class Main. class Solution is top-level package-private above Main.
CPP: #include <bits/stdc++.h>, using namespace std.
PYTHON: # USER_CODE_START / # USER_CODE_END markers.
JAVASCRIPT: top-level function, JSON.stringify for arrays.
C: no try-catch (platform handles RE). Arrays passed as (pointer, size).

Follow these example patterns (adapt names/types/values to the spec):

===HARNESS:JAVA===
import java.util.*;
// USER_CODE_START
class Solution { public int solve(int[] arr) { return 0; } }
// USER_CODE_END
public class Main {
    static void test(int[] arr, int expected, int tc, boolean hidden) {
        int got = new Solution().solve(arr);
        if (got == expected) System.out.println("TC:" + tc + ":PASS" + (hidden ? ":hidden" : ""));
        else if (hidden) System.out.println("TC:" + tc + ":FAIL:hidden");
        else System.out.println("TC:" + tc + ":FAIL:input=" + Arrays.toString(arr) + ":expected=" + expected + ":got=" + got);
    }
    public static void main(String[] a) {
        try { test(new int[]{1,2,3}, 6, 1, false); } catch (Exception e) { System.out.println("TC:1:FAIL:input=[1,2,3]:expected=6:got=ERR"); }
        try { test(new int[]{5}, 5, 2, false); } catch (Exception e) { System.out.println("TC:2:FAIL:input=[5]:expected=5:got=ERR"); }
    }
}
===HARNESS:CPP===
#include <bits/stdc++.h>
using namespace std;
// USER_CODE_START
class Solution { public: int solve(vector<int>& arr) { return 0; } };
// USER_CODE_END
void test(vector<int> arr, int expected, int tc, bool hidden=false) {
    Solution sol; int got = sol.solve(arr);
    if (got == expected) cout << "TC:" << tc << ":PASS" << (hidden ? ":hidden" : "") << "\n";
    else if (hidden) cout << "TC:" << tc << ":FAIL:hidden\n";
    else { cout << "TC:" << tc << ":FAIL:input=["; for(size_t i=0;i<arr.size();i++){if(i)cout<<",";cout<<arr[i];} cout << "]:expected=" << expected << ":got=" << got << "\n"; }
}
int main(){ try{test({1,2,3},6,1);}catch(...){cout<<"TC:1:FAIL:input=[1,2,3]:expected=6:got=ERR\n";} try{test({5},5,2);}catch(...){cout<<"TC:2:FAIL:input=[5]:expected=5:got=ERR\n";} return 0; }
===HARNESS:PYTHON===
# USER_CODE_START
class Solution: def solve(self, arr): return 0
# USER_CODE_END
def test(arr, expected, tc, hidden=False):
    got = Solution().solve(arr)
    if got == expected: print(f"TC:{tc}:PASS"+(":hidden" if hidden else ""))
    elif hidden: print(f"TC:{tc}:FAIL:hidden")
    else: print(f"TC:{tc}:FAIL:input={arr}:expected={expected}:got={got}")
try: test([1,2,3],6,1)
except: print("TC:1:FAIL:hidden")
try: test([5],5,2)
except: print("TC:2:FAIL:hidden")
===HARNESS:JAVASCRIPT===
// USER_CODE_START
function solve(arr) { return 0; }
// USER_CODE_END
function test(arr, expected, tc, hidden=false) {
    const got = solve(arr);
    if (got === expected) console.log(`TC:${tc}:PASS` + (hidden ? ':hidden' : ''));
    else if (hidden) console.log(`TC:${tc}:FAIL:hidden`);
    else console.log(`TC:${tc}:FAIL:input=[${arr}]:expected=${expected}:got=${got}`);
}
try { test([1,2,3],6,1); } catch(e) { console.log("TC:1:FAIL:hidden"); }
try { test([5],5,2); } catch(e) { console.log("TC:2:FAIL:hidden"); }
===HARNESS:C===
#include <stdio.h>
// USER_CODE_START
int solve(int* arr, int n) { return 0; }
// USER_CODE_END
void test(int* arr, int n, int expected, int tc, int hidden) {
    int got = solve(arr, n);
    if (got == expected) { if (hidden) printf("TC:%d:PASS:hidden\n",tc); else printf("TC:%d:PASS\n",tc); }
    else { if (hidden) printf("TC:%d:FAIL:hidden\n",tc);
    else { printf("TC:%d:FAIL:input=[",tc); for(int i=0;i<n;i++){if(i)printf(",");printf("%d",arr[i]);} printf("]:expected=%d:got=%d\n",expected,got); } }
}
int main(){ int t1[]={1,2,3}; test(t1,3,6,1,0); int t2[]={5}; test(t2,1,5,2,0); return 0; }"""


def call_ai(messages, max_tokens=8192, temperature=0.5):
    """Call NVIDIA NIM API."""
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": MODEL,
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": temperature,
        "top_p": 0.95,
        "stream": False
    }
    for attempt in range(5):
        try:
            r = requests.post(API_URL, json=payload, headers=headers, timeout=300)
            if r.status_code == 200:
                body = r.json()
                content = body["choices"][0]["message"]["content"]
                return content
            print(f"  API {r.status_code}: {r.text[:200]}")
            if r.status_code in (429, 503) and attempt < 4:
                time.sleep(8 * (attempt + 1))
                continue
            return None
        except Exception as e:
            print(f"  Network error: {e}")
            if attempt < 4:
                time.sleep(4 * (attempt + 1) * (attempt + 1) / 1000)
    return None


def generate_problem(query):
    """Pass 1: Generate problem spec JSON."""
    messages = [
        {"role": "system", "content": PASS1_SYSTEM},
        {"role": "user", "content": f"Design the problem spec for: {query}\n\nOutput ONLY the raw JSON spec object. Start with {{ and end with }}."}
    ]
    for attempt in range(4):
        temp = [0.5, 0.65, 0.8, 0.9][attempt]
        raw = call_ai(messages, temperature=temp)
        if not raw:
            continue
        # Extract JSON from response
        raw = raw.strip()
        if raw.startswith("```"):
            raw = re.sub(r"^```\w*\n?", "", raw)
            raw = re.sub(r"\n?```$", "", raw)
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            print(f"  Pass 1 attempt {attempt+1}: JSON parse failed, retrying...")
    return None


def generate_harnesses(spec):
    """Pass 2: Generate 5 harnesses."""
    # Build compact spec for harness generation
    slim = {
        "title": spec["problem"]["title"],
        "signature": spec["signature"],
        "tests": spec["tests"]
    }
    spec_str = json.dumps(slim)
    
    messages = [
        {"role": "system", "content": HARNESS_SYSTEM},
        {"role": "user", "content": f"Spec:\n{spec_str}\n\nOutput all five harnesses. Before each one put a line exactly like ===HARNESS:JAVA=== (then CPP, PYTHON, JAVASCRIPT, C). Raw source only, no markdown fences, no commentary."}
    ]
    for attempt in range(3):
        temp = [0.5, 0.65, 0.8][attempt]
        raw = call_ai(messages, temperature=temp)
        if not raw:
            continue
        # Parse harnesses
        harnesses = {}
        current_lang = None
        current_code = []
        for line in raw.split("\n"):
            m = re.match(r"===HARNESS:(JAVA|CPP|PYTHON|JAVASCRIPT|C)===", line.strip())
            if m:
                if current_lang and current_code:
                    harnesses[current_lang] = "\n".join(current_code).strip()
                current_lang = m.group(1)
                current_code = []
            elif current_lang:
                current_code.append(line)
        if current_lang and current_code:
            harnesses[current_lang] = "\n".join(current_code).strip()
        
        if set(harnesses.keys()) == {"JAVA", "CPP", "PYTHON", "JAVASCRIPT", "C"}:
            return harnesses
        print(f"  Pass 2 attempt {attempt+1}: missing languages, got {set(harnesses.keys())}")
    return None


def save_to_db(problem, harnesses):
    """Insert problem + snippets directly into PostgreSQL."""
    import psycopg2
    conn = psycopg2.connect(host=DB_HOST, port=DB_PORT, dbname=DB_NAME, user=DB_USER, password=DB_PASS)
    cur = conn.cursor()
    
    p = problem
    cur.execute("""
        INSERT INTO problems (title, description, input_format, output_format, constraints, 
            time_limit, memory_limit, level, active, topics, example1, example2, example3)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (
        p["title"], p["description"], p.get("inputFormat", ""), p.get("outputFormat", ""),
        p.get("constraints", ""), p["timeLimit"], p["memoryLimit"], p["level"],
        True, p.get("topics", ""), p.get("example1", ""), p.get("example2", ""), p.get("example3", "")
    ))
    pid = cur.fetchone()[0]
    
    for lang, code in harnesses.items():
        cur.execute("""
            INSERT INTO code_snippets (problem_id, language, solution_template, created_at, updated_at)
            VALUES (%s, %s, %s, NOW(), NOW())
        """, (pid, lang, code))
    
    conn.commit()
    cur.close()
    conn.close()
    return pid


def batch_generate(start=0, count=None):
    """Main batch loop."""
    wb = openpyxl.load_workbook(SHEET_PATH)
    ws = wb.active
    
    questions = []
    for row in range(2, ws.max_row + 1):
        name = str(ws.cell(row, 4).value or "").strip()
        diff = str(ws.cell(row, 5).value or "").strip()
        if not name or name == "None":
            continue
        questions.append((name, diff))
    
    if count:
        questions = questions[start:start + count]
    elif start > 0:
        questions = questions[start:]
    
    print(f"Total questions to generate: {len(questions)}")
    print(f"Model: {MODEL}")
    print("=" * 60)
    
    success, failed = 0, 0
    for i, (name, difficulty) in enumerate(questions, 1):
        print(f"\n[{i}/{len(questions)}] {name} ({difficulty})")
        
        # Pass 1: Generate problem spec
        print("  Pass 1: Generating problem spec...")
        spec = generate_problem(name)
        if not spec or "problem" not in spec:
            print(f"  FAILED: Could not generate problem spec")
            failed += 1
            continue
        
        p = spec["problem"]
        print(f"  Generated: {p['title']} ({p['level']}) — {p.get('topics','')}")
        
        # Verify reference solution
        ref_sol = spec.get("referenceSolutionPython", "")
        if ref_sol:
            verified = verify_reference(ref_sol, spec)
            if verified:
                print(f"  Reference solution: VERIFIED")
            else:
                print(f"  Reference solution: FAILED verification (using AI expected values)")
        
        # Pass 2: Generate harnesses
        print("  Pass 2: Generating harnesses...")
        harnesses = generate_harnesses(spec)
        if not harnesses:
            print(f"  FAILED: Could not generate harnesses")
            failed += 1
            continue
        
        # Save to DB
        try:
            pid = save_to_db(p, harnesses)
            print(f"  SAVED: pid={pid} ({len(harnesses)} harnesses)")
            success += 1
        except Exception as e:
            print(f"  DB ERROR: {e}")
            failed += 1
        
        # Rate limit delay between problems
        if i < len(questions):
            time.sleep(3)
    
    print(f"\n{'=' * 60}")
    print(f"DONE. Success: {success}, Failed: {failed}")


def verify_reference(ref_sol, spec):
    """Execute reference Python solution against test inputs to verify."""
    try:
        sig = spec["signature"]
        tests = spec["tests"]
        func_name = sig["name"]
        
        # Build execution script
        code = ref_sol + "\n\n"
        for tc in tests:
            args_str = ", ".join([
                json.dumps(tc["args"][p["name"]])
                for p in sig["params"]
            ])
            code += f"result = {func_name}({args_str})\n"
            code += f"assert result == {json.dumps(tc['expected'])}, f'Test {{tc}} failed: {{result}} != {json.dumps(tc['expected'])}'\n"
        
        with tempfile.NamedTemporaryFile(mode="w", suffix=".py", delete=False) as f:
            f.write(code)
            tmp = f.name
        
        r = subprocess.run(["python3", tmp], capture_output=True, text=True, timeout=15)
        os.unlink(tmp)
        return r.returncode == 0
    except Exception as e:
        print(f"  Verify error: {e}")
        return False


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Batch generate CodeCombat problems from Excel sheet")
    ap.add_argument("--start", type=int, default=0, help="Start index (0-based)")
    ap.add_argument("--count", type=int, default=0, help="Number to generate (0=all)")
    ap.add_argument("--dry-run", action="store_true", help="Print questions without generating")
    args = ap.parse_args()
    
    wb = openpyxl.load_workbook(SHEET_PATH)
    ws = wb.active
    questions = []
    for row in range(2, ws.max_row + 1):
        name = str(ws.cell(row, 4).value or "").strip()
        if name and name != "None":
            questions.append(name)
    
    if args.dry_run:
        for i, q in enumerate(questions):
            print(f"{i+1}. {q}")
        print(f"\nTotal: {len(questions)}")
        sys.exit(0)
    
    batch_generate(start=args.start, count=args.count)
